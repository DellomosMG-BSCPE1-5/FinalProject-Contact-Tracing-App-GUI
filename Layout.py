#This .py file contains the code for the layout of the interface
#import the tkinter module
import pathlib
import openpyxl
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from openpyxl import Workbook
from datetime import datetime as dt

#creation of excel file
def data_file():
    data_file = pathlib.Path('Contact Tracing Data.xlsx')
    #checking if the data_file exists
    if data_file.exists():
        pass
    #if the file already exist, append the following for header
    else:
        data_file = Workbook()
        active_sheet = data_file.active
        active_sheet['A1'] = "First Name"
        active_sheet['B1'] = "Last Name"
        active_sheet['C1'] = "Contact Number"
        active_sheet['D1'] = "Email Address"
        active_sheet['E1'] = "Residential Address"
        active_sheet['F1'] = "Entry Time"
        active_sheet['G1'] = "Entry Date"
        active_sheet['H1'] = "Vaccine Received"
        active_sheet['I1'] = "COVID Symptoms"
        data_file.save('Contact Tracing Data.xlsx')

#save_the_entries method for the command of the submit button
def save_the_entries():
    #getting the contact information and additional data
    first_name = first_name_input.get()
    last_name = last_name_input.get()
    contact_number = contact_num_input.get()
    email_address = email_address_input.get()
    residential_address = residential_address_input.get()
    entry_date = entry_date_input.get()
    entry_time = [entry_time_input.get()]
    if entry_time_am_pm.get() == "A.M.":
        entry_time.append("A.M.")
    if entry_time_am_pm.get() == "P.M.":
        entry_time.append("P.M.")
    entry_time_str = " ".join(entry_time)

    #input validation
    if first_name == "" or last_name == "" or contact_number == "" or residential_address == "" or entry_date == "" or entry_time == "":
        messagebox.showwarning(title="WARNING!", message="All required details must be filled.")
    elif len(contact_number) != 11:
        messagebox.showwarning(title="Invalid Contact Number", message="Invalid Input. Contact number must be 11 digits.")
    else:
        #moving the data into excel file
        data_file=openpyxl.load_workbook('Contact Tracing Data.xlsx')
        active_sheet = data_file.active
        active_sheet.cell(column=1, row=active_sheet.max_row+1, value=first_name)
        active_sheet.cell(column=2, row=active_sheet.max_row, value=last_name)
        active_sheet.cell(column=3, row=active_sheet.max_row, value=contact_number)
        active_sheet.cell(column=4, row=active_sheet.max_row, value=email_address)
        active_sheet.cell(column=5, row=active_sheet.max_row, value=residential_address)
        active_sheet.cell(column=6, row=active_sheet.max_row, value=entry_time_str)
        active_sheet.cell(column=7, row=active_sheet.max_row, value=entry_date)    
        #getting the value from radiobutton then move into the excel file
        if vaccines_received_var.get() == 1:
            active_sheet.cell(column=8, row=active_sheet.max_row, value="None")
        if vaccines_received_var.get() == 2:
            active_sheet.cell(column=8, row=active_sheet.max_row, value="First Dose")
        if vaccines_received_var.get() == 3:
            active_sheet.cell(column=8, row=active_sheet.max_row, value="Second Dose")
        if vaccines_received_var.get() == 4:
            active_sheet.cell(column=8, row=active_sheet.max_row, value="First Booster Shot")
        if vaccines_received_var.get() == 5:
            active_sheet.cell(column=8, row=active_sheet.max_row, value="Second Booster Shot")    
        #getting the values from checkbuttons then move in the excel file
        symptoms = []
        if none_of_the_above_var.get() == 1:
            symptoms.append("No")
        if fever_var.get() == 1:
            symptoms.append("Fever")
        if cough_var.get() == 1:
            symptoms.append("Cough")
        if sore_throat_var.get() == 1:
            symptoms.append("Sore Throat")
        if lss_of_tst_smll_var.get() == 1:
            symptoms.append("Loss of Taste and Smell")
        if mscls_bdy_pains_var.get() == 1:
            symptoms.append("Muscles and Body Pains")
        symptoms_str = ", ".join(symptoms)
        active_sheet.cell(column=9, row=active_sheet.max_row, value=symptoms_str)

        #save the entries    
        data_file.save('Contact Tracing Data.xlsx')

        #messagebox to know that the user successfully submitted his/her entry
        messagebox.showinfo(title="Information Submitted", message="Entries are Submitted Successfully")

        #clear the all of the fields after submission
        first_name_input.delete(0, END)
        last_name_input.delete(0, END)
        contact_num_input.delete(0, END)
        email_address_input.delete(0, END)
        residential_address_input.delete(0, END)
        entry_time_input.delete(0, END)
        entry_date_input.delete(0, END)
        none_of_the_above.deselect()
        fever.deselect()
        cough.deselect()
        sore_throat.deselect()
        lss_of_tst_smll.deselect()
        mscls_bdy_pains.deselect()
        vaccines_received_var.set(0)

#searching through excel file    
def search_records():
    #get the search data from the search bar
    search_data = search.get()
    data_file=openpyxl.load_workbook('Contact Tracing Data.xlsx')
    active_sheet = data_file.active
    #create an empty list to store the found data
    found_records = []
    # Check if the search data exists in any cells then if a match is found, append the entire row to the found_records list
    for row in active_sheet.iter_rows(min_row=2, values_only=True):
        if search_data.lower() in ' '.join(str(cell).lower() for cell in row):
            found_records.append(row)
    # If there are matching records found, create a string containing the data listed in found_record list
    if found_records:
        search_result = "\n\n".join("\n".join(f"{label}: {value}" for label, value in zip(
            ["First Name", "Last Name", "Contact Number", "Email Address", "Residential Address", 
             "Entry Time", "Entry Date", "Vaccine Received", "COVID Symptoms"], row)) for row in found_records)
        #create a new frame to display searched entries
        display_searched_result_frame = LabelFrame(window, bg="#F5F5F5", bd=0)
        display_searched_result_frame.place(x=20, y=75, width=470, height=680)
        #create a text widget to show the search results
        search_result_text = Text(display_searched_result_frame, font=('consolas', 12), fg="black", bg="#F5F5F5", wrap="word")
        search_result_text.pack(fill=BOTH, expand=True)
        search_result_text.insert(END, search_result)
        #create a scrollbar for the text widget
        scrollbar = Scrollbar(display_searched_result_frame, command=search_result_text.yview)
        scrollbar.pack(side=RIGHT)

        search_result_text.config(yscrollcommand=scrollbar.set)
        
        # Create a "Close" button to close the search results frame
        close_button = Button(display_searched_result_frame, text="Close", bg="red3", fg="white", font=('consolas', 20, "bold"), relief=GROOVE, bd=3, command=display_searched_result_frame.destroy)
        close_button.place(x=0, y=640, width=470, height=43)
    #if nothing matches the searched word, a message box will appear to inform the user
    else:
        messagebox.showinfo(title="Search Result", message="No matching records found.")
    
    #clear the search bar input after displaying the results or showing the message
    search.set("") 

#create the main window
window = Tk()
window.title("Safe Trace")
window.iconbitmap("C:\\Users\\Mary Grace\\Desktop\\Object-Oriented Programming\\FinalProject-Contact-Tracing-App-GUI\\FinalProject-Contact-Tracing-App-GUI\\safetrace.ico")
window.geometry("1920x1200")
window.configure(bg="#F5F5F5")

#create the frames
header_frame=Label(window,pady=2,bd=12,bg="#C42300", justify=CENTER) #top frame
header_frame.pack(fill=X)
left_frame = LabelFrame(window, bg="#F5F5F5", borderwidth=0) #left frame
left_frame.place(x=20, y=60, width=470, height=730)
app_title= Label(left_frame, text='Safe Trace', font=('Couture',35, 'bold'), bg="#F5F5F5", fg='red3') #left frame: title
app_title.grid(row=0, column=0, padx=75, pady=7)
poster = PhotoImage(file = 'C:\\Users\\Mary Grace\\Desktop\\Object-Oriented Programming\\FinalProject-Contact-Tracing-App-GUI\\FinalProject-Contact-Tracing-App-GUI\\OOP POSTER.png') #left frame: poster
Label(left_frame, image = poster).place(x=5, y=73)

#right frame
contact_info_frame = LabelFrame(window,text='Contact Information', font=('consolas', 15), fg= "black", bg="#F5F5F5")
addtl_info_frame = LabelFrame(window,bg="#F5F5F5")
covid_rltd_frame_1 = LabelFrame(window,bg="#F5F5F5")
covid_rltd_frame_2 = LabelFrame(window,bg="#F5F5F5")

contact_info_frame.place(x=500, y=125, width=1015, height=270)
addtl_info_frame.place(x=500, y=405, width=1015, height=65)
covid_rltd_frame_1.place(x=1013, y=480, width=502, height=200)
covid_rltd_frame_2.place(x=500, y=480, width=502, height=277)

#right frame; Contact Information; LABELS
first_name_label = Label(contact_info_frame, text="First Name      ", font=('consolas', 13), fg= "dark red", bg="#F5F5F5")
last_name_label = Label(contact_info_frame, text="  Last Name", font=('consolas', 13), fg= "dark red", bg="#F5F5F5")
contact_num_label = Label(contact_info_frame, text="   Contact Number     ", font=('consolas', 13), fg= "dark red", bg="#F5F5F5")
email_address_label = Label(contact_info_frame, text="Email Address (Optional)", font=('consolas', 13), fg= "dark red", bg="#F5F5F5")
residential_address_label = Label(contact_info_frame, text="   Residential Address", font=('consolas', 13), fg= "dark red", bg="#F5F5F5")
vaccines_received_label = Label(covid_rltd_frame_1, text="Received  COVID-19 Vaccine:", font=('consolas', 13), fg= "dark red", bg="#F5F5F5")
covid_symptoms_label = Label(covid_rltd_frame_2, text="In the past week, have you experienced any of the \n symptoms listed below?", font=('consolas', 13), fg= "dark red", bg="#F5F5F5")

first_name_label.grid(row=0, column=0, padx=55, pady=10)
last_name_label.grid(row=0, column=1, padx=0, pady=10)
contact_num_label.grid(row=1, column=0, padx=25, pady=40)
email_address_label.grid(row=1, column=1, padx=187, pady=40)
residential_address_label.grid(row=2, column=0, padx=0, pady=10)
vaccines_received_label.pack(anchor="w", padx=15, pady=5)
covid_symptoms_label.pack(anchor="w", padx=21, pady=5)

#right frame; Contact Information; ENTRY
first_name_input = Entry(contact_info_frame, font=('consolas', 12), fg= "black", bg="#E6E4E4",relief=GROOVE, bd=2)
last_name_input = Entry(contact_info_frame, font=('consolas', 12), fg= "black", bg="#E6E4E4",relief=GROOVE, bd=2 )
contact_num_input = Entry(contact_info_frame, font=('consolas', 12), fg= "black", bg="#E6E4E4",relief=GROOVE, bd=2)
email_address_input = Entry(contact_info_frame, font=('consolas', 12), fg= "black", bg="#E6E4E4",relief=GROOVE, bd=2)
residential_address_input = Entry(contact_info_frame, font=('consolas', 12), fg= "black", bg="#E6E4E4",relief=GROOVE, bd=2)

first_name_input.place(x=60, y=40, width=425, height=30)
last_name_input.place(x=525, y=40, width=425, height=30)
contact_num_input.place(x=60, y=115, width=350, height=30)
email_address_input.place(x=450, y=115, width=500, height=30) 
residential_address_input.place(x=60, y=190, width=890, height=30)

#right frame; Additional Information; LABELS
entry_time_label = Label(addtl_info_frame, text="Time of Arrival at the Premises:", font=('consolas', 13), fg= "dark red", bg="#F5F5F5")
entry_date_label = Label(addtl_info_frame, text="Date of Arrival at the Premises:", font=('consolas', 13), fg= "dark red", bg="#F5F5F5")

entry_time_label.grid(row=0, column=0, padx=55, pady=15)
entry_date_label.grid(row=0, column=1, padx=125, pady=15)

#right frame; Additional Information
#time
entry_time_input = Entry(addtl_info_frame, font=('consolas', 12), fg= "black", bg="#E6E4E4",relief=GROOVE, bd=2)
entry_time_am_pm = ttk.Combobox(addtl_info_frame, values=["A.M.", "P.M."], font=("consolas", 11))
entry_time_input.place(x=350, y=15, width=70, height=30)
entry_time_am_pm.place(x=423, y=14.3, width=50, height=30)

#date
entry_date_var = StringVar()
date_today = dt.now()
entry_date = date_today.strftime("%d/%m/%Y")
entry_date_input = Entry(addtl_info_frame, textvariable=entry_date_var, font=('consolas', 12), fg= "black", bg="#E6E4E4",relief=GROOVE, bd=2)
entry_date_input.place(x=825, y=15, width=125, height=30)
entry_date_var.set(entry_date)

#Checkbox Symptoms
none_of_the_above_var = IntVar()
fever_var = IntVar()
cough_var = IntVar()
sore_throat_var = IntVar()
lss_of_tst_smll_var = IntVar()
mscls_bdy_pains_var = IntVar()

none_of_the_above = Checkbutton(covid_rltd_frame_2, text ='NO', variable = none_of_the_above_var, bg="#F5F5F5", font=('consolas', 12), onvalue=1, offvalue=0)
fever = Checkbutton(covid_rltd_frame_2, text ='Fever', variable = fever_var, bg="#F5F5F5", font=('consolas', 12), onvalue=1, offvalue=0)
cough = Checkbutton(covid_rltd_frame_2, text ='Cough', variable = cough_var, bg="#F5F5F5", font=('consolas', 12), onvalue=1, offvalue=0)
sore_throat = Checkbutton(covid_rltd_frame_2, text ='Sore Throat', variable = sore_throat_var, bg="#F5F5F5", font=('consolas', 12), onvalue=1, offvalue=0)
lss_of_tst_smll = Checkbutton(covid_rltd_frame_2, text ='Loss of Taste and Smell', variable = lss_of_tst_smll_var, bg="#F5F5F5", font=('consolas', 12), onvalue=1, offvalue=0)
mscls_bdy_pains = Checkbutton(covid_rltd_frame_2, text ='Muscle and Body Pains', variable = mscls_bdy_pains_var, bg="#F5F5F5", font=('consolas', 12), onvalue=1, offvalue=0)

none_of_the_above.pack(anchor="w", padx=21, pady=0)
fever.pack(anchor="w", padx=21, pady=0)
cough.pack(anchor="w", padx=21, pady=0)
sore_throat.pack(anchor="w", padx=21, pady=0)
lss_of_tst_smll.pack(anchor="w", padx=21, pady=0)
mscls_bdy_pains.pack(anchor="w", padx=21, pady=0)

#create the necessary buttons
#radiobuttons
vaccines_received_var = IntVar(covid_rltd_frame_1, "0")
vaccines_received_values = {"None" : "1", "First Dose" : "2", "Second Dose" : "3", "First Booster Shot" : "4", "Second Booster Shot" : "5"}
for (text, value) in vaccines_received_values.items():
    vaccines_received = Radiobutton(covid_rltd_frame_1, text = text, variable = vaccines_received_var, value = value, bg="#F5F5F5", font=('consolas', 12))
    vaccines_received.pack(anchor="w", padx=15, pady=0)

#search bar and search button
search=StringVar()
search_bar_input = Entry(window, textvariable=search, font=('consolas', 12), fg= "black", bg="#E6E4E4",relief=GROOVE, bd=2)
search_bar_input.place(x=500, y=75, width=775, height=45)
search_button = Button(window, text="Search", compound=LEFT, width=200, bg='red3', fg="white", font=('consolas', 20, 'bold'), relief=GROOVE, bd=3, command=search_records)
search_button.place(x=1285, y=75, width=230, height=45)

#submit button
submit_button = Button(window,text="Submit", width="32", height="1", bg="red3", fg="white", font=("consolas", 20, "bold"), relief=GROOVE, bd=3, command=save_the_entries)
submit_button.place(x=1013, y=690, width=502, height=65)



window.mainloop()